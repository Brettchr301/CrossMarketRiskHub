from __future__ import annotations

from io import BytesIO

from fastapi import APIRouter, Depends, HTTPException
from fastapi.responses import StreamingResponse
from openpyxl import Workbook
from sqlalchemy.orm import Session

from app.api.narrative import build_narrative
from app.modeling.analytics import CorrelationAnalyticsService
from app.modeling.global_scan import GlobalOpportunityService
from app.modeling.research_hub import ResearchHubService
from app.db.repositories import (
    latest_backtest_metrics,
    latest_commodity_distributions,
    latest_equity_price,
    latest_event_probabilities,
    latest_fundamental_state,
    latest_options_distribution,
    latest_signals,
    latest_valuation,
)
from app.db.session import get_db
from app.pipeline.orchestration import PipelineOrchestrator
from app.schemas import (
    BacktestMetrics,
    CommodityDistribution,
    CorrelationSnapshot,
    PredictiveContractsSnapshot,
    PredictiveContract,
    TickerResearchView,
    ResearchSeriesPoint,
    ModelValidationStats,
    ShippingHedgeStats,
    EventProbability,
    FundamentalState,
    ImpliedDistribution,
    Signal,
    TickerCorrelationStats,
    CorrelationDriver,
    CommodityTypeEffectiveness,
    ValuationDistribution,
    GlobalOpportunity,
    GlobalOpportunitiesSnapshot,
)


router = APIRouter(prefix="/v1", tags=["valuation-platform"])
orchestrator = PipelineOrchestrator()
analytics = CorrelationAnalyticsService()
research_hub = ResearchHubService()
global_scan = GlobalOpportunityService()


def _ensure_seed_data(db: Session) -> None:
    if latest_signals(db, limit=1):
        return
    orchestrator.run_daily(db)


@router.get("/events/probabilities", response_model=list[EventProbability])
def get_event_probabilities(db: Session = Depends(get_db)):
    _ensure_seed_data(db)
    rows = latest_event_probabilities(db)
    return [
        EventProbability(
            event_id=row.event_id,
            prob=row.prob,
            ci_low=row.ci_low,
            ci_high=row.ci_high,
            as_of=row.as_of,
        )
        for row in rows
    ]


@router.get("/commodities/distributions", response_model=list[CommodityDistribution])
def get_commodity_distributions(db: Session = Depends(get_db)):
    _ensure_seed_data(db)
    rows = latest_commodity_distributions(db)
    return [
        CommodityDistribution(
            symbol=row.symbol,
            horizon_days=row.horizon_days,
            p05=row.p05,
            p50=row.p50,
            p95=row.p95,
            as_of=row.as_of,
        )
        for row in rows
    ]


@router.get("/companies/{ticker}/fundamental-state", response_model=FundamentalState)
def get_company_fundamental_state(ticker: str, db: Session = Depends(get_db)):
    _ensure_seed_data(db)
    row = latest_fundamental_state(db, ticker=ticker)
    if row is None:
        raise HTTPException(status_code=404, detail=f"Fundamental state not found for ticker {ticker}")
    return FundamentalState(
        ticker=row.ticker,
        guidance_period=row.guidance_period,
        production=row.production,
        costs=row.cost_per_unit + row.transport_cost + row.sga,
        capex=row.capex,
        debt=row.debt,
        shares=row.share_count,
        confidence=row.confidence,
        as_of=row.as_of,
    )


@router.get("/companies/{ticker}/valuation-distribution", response_model=ValuationDistribution)
def get_company_valuation_distribution(ticker: str, db: Session = Depends(get_db)):
    _ensure_seed_data(db)
    row = latest_valuation(db, ticker=ticker)
    if row is None:
        raise HTTPException(status_code=404, detail=f"Valuation distribution not found for ticker {ticker}")
    return ValuationDistribution(
        ticker=row.ticker,
        ev_p50=row.ev_p50,
        equity_ps_p50=row.equity_ps_p50,
        expected_return_net_cost=row.expected_return_net_cost,
        downside_p05=row.downside_p05,
        ev_p05=row.ev_p05,
        ev_p95=row.ev_p95,
        equity_ps_p05=row.equity_ps_p05,
        equity_ps_p95=row.equity_ps_p95,
        as_of=row.as_of,
    )


@router.get("/companies/{ticker}/valuation-distribution.xlsx")
def export_company_valuation_distribution_xlsx(ticker: str, db: Session = Depends(get_db)):
    _ensure_seed_data(db)
    val = latest_valuation(db, ticker=ticker)
    fs = latest_fundamental_state(db, ticker=ticker)
    if val is None:
        raise HTTPException(status_code=404, detail=f"Valuation distribution not found for ticker {ticker}")
    if fs is None:
        raise HTTPException(status_code=404, detail=f"Fundamental state not found for ticker {ticker}")
    spot = latest_equity_price(db, ticker=ticker)
    if spot is None:
        raise HTTPException(status_code=404, detail=f"Spot price not found for ticker {ticker}")

    wb = Workbook()
    ws = wb.active
    ws.title = "valuation_summary"
    ws.append(["Ticker", val.ticker])
    ws.append(["As of", str(val.as_of)])
    ws.append(["Spot Price", float(spot)])
    ws.append([])
    ws.append(["Metric", "Value"])
    ws.append(["EV P05", float(val.ev_p05)])
    ws.append(["EV P50", float(val.ev_p50)])
    ws.append(["EV P95", float(val.ev_p95)])
    ws.append(["Equity/Share P05", float(val.equity_ps_p05)])
    ws.append(["Equity/Share P50", float(val.equity_ps_p50)])
    ws.append(["Equity/Share P95", float(val.equity_ps_p95)])
    ws.append(["Expected Return Net Cost", float(val.expected_return_net_cost)])
    ws.append(["Downside P05", float(val.downside_p05)])

    ws2 = wb.create_sheet(title="fundamentals")
    ws2.append(["Field", "Value"])
    ws2.append(["Guidance Period", fs.guidance_period])
    ws2.append(["Production", float(fs.production)])
    ws2.append(["Costs (aggregate)", float(fs.cost_per_unit + fs.transport_cost + fs.sga)])
    ws2.append(["Capex", float(fs.capex)])
    ws2.append(["Debt", float(fs.debt)])
    ws2.append(["Interest Rate", float(fs.interest_rate)])
    ws2.append(["Hedge Ratio", float(fs.hedge_ratio)])
    ws2.append(["Utilization", float(fs.utilization)])
    ws2.append(["Share Count", float(fs.share_count)])
    ws2.append(["Confidence", float(fs.confidence)])
    ws2.append(["Meta Payload", str(fs.meta_payload)])

    ws4 = wb.create_sheet(title="drivers")
    ws4.append(["Driver", "Value"])
    meta = fs.meta_payload or {}
    for key in sorted(meta):
        value = meta[key]
        if isinstance(value, (int, float, str)):
            ws4.append([key, value])
    for row in latest_event_probabilities(db):
        ws4.append([f"event_prob::{row.event_id}", float(row.prob)])

    ws3 = wb.create_sheet(title="check_math")
    ws3.append(["Check", "Formula/Value"])
    ws3.append(["Spot Price", float(spot)])
    ws3.append(["Equity/Share P50", float(val.equity_ps_p50)])
    ws3.append(["Implied Gross Return", "=(B3-B2)/B2"])
    ws3.append(["Expected Return Net Cost (API)", float(val.expected_return_net_cost)])
    ws3.append(["Difference (API - Formula)", "=B5-B4"])
    ws3.append(["EV P50", float(val.ev_p50)])
    ws3.append(["EV P95", float(val.ev_p95)])
    ws3.append(["EV Range %", "=(B8-B7)/B7"])

    output = BytesIO()
    wb.save(output)
    output.seek(0)
    file_name = f"{ticker.upper()}_valuation_distribution.xlsx"
    headers = {"Content-Disposition": f'attachment; filename="{file_name}"'}
    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers=headers,
    )


@router.get("/options/{ticker}/implied-distribution", response_model=ImpliedDistribution)
def get_options_implied_distribution(ticker: str, db: Session = Depends(get_db)):
    _ensure_seed_data(db)
    row = latest_options_distribution(db, ticker=ticker)
    if row is None:
        raise HTTPException(status_code=404, detail=f"Options implied distribution not found for ticker {ticker}")
    return ImpliedDistribution(
        ticker=row.ticker,
        horizon_days=row.horizon_days,
        mean_return=row.mean_return,
        std_return=row.std_return,
        downside_p05=row.downside_p05,
        upside_p95=row.upside_p95,
        as_of=row.as_of,
    )


@router.get("/signals", response_model=list[Signal])
def get_signals(db: Session = Depends(get_db)):
    _ensure_seed_data(db)
    rows = latest_signals(db, limit=50)
    return [
        Signal(
            ticker=row.ticker,
            score=row.score,
            direction=row.direction,
            holding_period_days=row.holding_period_days,
            expected_return_net_cost=row.expected_return_net_cost,
            risk_flags=[x for x in row.risk_flags.split(",") if x],
            as_of=row.as_of,
        )
        for row in rows
    ]


@router.get("/backtest/metrics", response_model=BacktestMetrics)
def get_backtest_metrics(db: Session = Depends(get_db)):
    _ensure_seed_data(db)
    row = latest_backtest_metrics(db)
    if row is None:
        orchestrator.run_daily(db)
        row = latest_backtest_metrics(db)
        if row is None:
            raise HTTPException(status_code=404, detail="Backtest metrics unavailable")
    return BacktestMetrics(
        sharpe=row.sharpe,
        hit_rate=row.hit_rate,
        average_alpha=row.average_alpha,
        max_drawdown=row.max_drawdown,
        turnover=row.turnover,
        capacity=row.capacity,
        irr=row.irr,
        as_of=row.as_of,
    )


@router.get("/analytics/correlations", response_model=CorrelationSnapshot)
def get_correlation_snapshot(lookback_days: int = 260):
    lookback_days = max(90, min(1200, int(lookback_days)))
    try:
        snap = analytics.build_snapshot(lookback_days=lookback_days)
    except Exception as exc:
        raise HTTPException(status_code=503, detail=f"Correlation snapshot unavailable: {exc}") from exc

    return CorrelationSnapshot(
        as_of=snap.as_of,
        lookback_days=snap.lookback_days,
        tickers=[
            TickerCorrelationStats(
                ticker=row["ticker"],
                sample_size=row["sample_size"],
                corr_brent=row.get("corr_brent"),
                corr_wti=row.get("corr_wti"),
                corr_shipping=row.get("corr_shipping"),
                top_drivers=[
                    CorrelationDriver(
                        name=d["name"],
                        source=d["source"],
                        correlation=d["correlation"],
                        lag_days=d["lag_days"],
                    )
                    for d in row.get("top_drivers", [])
                ],
            )
            for row in snap.tickers
        ],
    )


@router.get("/analytics/predictive-contracts", response_model=PredictiveContractsSnapshot)
def get_predictive_contracts_snapshot(lookback_days: int = 420):
    lookback_days = max(180, min(1500, int(lookback_days)))
    try:
        snap = research_hub.predictive_contracts_snapshot_async(lookback_days=lookback_days)
    except Exception as exc:
        raise HTTPException(status_code=503, detail=f"Predictive-contract snapshot unavailable: {exc}") from exc
    if snap is None:
        raise HTTPException(status_code=503, detail="computing: predictive contracts in progress - retry in 2-3 minutes")
    return PredictiveContractsSnapshot(
        as_of=snap["as_of"],
        lookback_days=snap["lookback_days"],
        contracts=[
            PredictiveContract(
                market_id=row["market_id"],
                question=row["question"],
                category=row["category"],
                best_target=row["best_target"],
                lead_days=row["lead_days"],
                correlation=row["correlation"],
                liquidity_score=row["liquidity_score"],
                staleness_days=row["staleness_days"],
                predictive_score=row["predictive_score"],
            )
            for row in snap["contracts"]
        ],
    )


@router.get("/analytics/tickers/{ticker}/research", response_model=TickerResearchView)
def get_ticker_research_view(ticker: str, lookback_days: int = 260):
    lookback_days = max(120, min(1200, int(lookback_days)))
    try:
        snap = research_hub.ticker_research_async(ticker=ticker, lookback_days=lookback_days)
    except Exception as exc:
        raise HTTPException(status_code=503, detail=f"Ticker research unavailable for {ticker}: {exc}") from exc
    if snap is None:
        raise HTTPException(status_code=503, detail=f"computing: ticker research for {ticker} in progress - retry in 2-3 minutes")
    return TickerResearchView(
        ticker=snap["ticker"],
        as_of=snap["as_of"],
        series=[
            ResearchSeriesPoint(
                date=row["date"],
                stock=row["stock"],
                brent=row["brent"],
                wti=row["wti"],
                shipping_spot=row["shipping_spot"],
                shipping_fwd=row["shipping_fwd"],
                event_hormuz=row["event_hormuz"],
                event_red_sea=row["event_red_sea"],
                event_oil_100=row["event_oil_100"],
            )
            for row in snap["series"]
        ],
        top_predictive_contracts=[
            PredictiveContract(
                market_id=row["market_id"],
                question=row["question"],
                category=row["category"],
                best_target=row["best_target"],
                lead_days=row["lead_days"],
                correlation=row["correlation"],
                liquidity_score=row["liquidity_score"],
                staleness_days=row["staleness_days"],
                predictive_score=row["predictive_score"],
            )
            for row in snap["top_predictive_contracts"]
        ],
        validation=ModelValidationStats(
            baseline_hit_rate=snap["validation"]["baseline_hit_rate"],
            enriched_hit_rate=snap["validation"]["enriched_hit_rate"],
            baseline_mae=snap["validation"]["baseline_mae"],
            enriched_mae=snap["validation"]["enriched_mae"],
            enriched_expected_return_20d=snap["validation"]["enriched_expected_return_20d"],
            fair_value_price=snap["validation"]["fair_value_price"],
            spot_price=snap["validation"]["spot_price"],
        ),
        hedge=ShippingHedgeStats(
            spot_proxy=snap["hedge"]["spot_proxy"],
            forward_proxy=snap["hedge"]["forward_proxy"],
            current_basis_pct=snap["hedge"]["current_basis_pct"],
            one_month_expected_basis_pct=snap["hedge"]["one_month_expected_basis_pct"],
            hedge_beta_to_forward=snap["hedge"]["hedge_beta_to_forward"],
        ),
    )


@router.get("/analytics/global-opportunities", response_model=GlobalOpportunitiesSnapshot)
def get_global_opportunities(
    lookback_days: int = 780,
    min_modeled_count: int = 200,
    max_rows: int = 220,
):
    lookback_days = max(360, min(2200, int(lookback_days)))
    min_modeled_count = max(80, min(500, int(min_modeled_count)))
    max_rows = max(50, min(500, int(max_rows)))
    try:
        snap = global_scan.scan_async(
            lookback_days=lookback_days,
            min_modeled_count=min_modeled_count,
            max_rows=max_rows,
        )
    except Exception as exc:
        raise HTTPException(status_code=503, detail=f"Global opportunity scan unavailable: {exc}") from exc
    if snap is None:
        raise HTTPException(status_code=503, detail="computing: global scan in progress - retry in 2-3 minutes")
    return GlobalOpportunitiesSnapshot(
        as_of=snap["as_of"],
        lookback_days=snap["lookback_days"],
        universe_size=snap["universe_size"],
        modeled_count=snap["modeled_count"],
        spot_proxy=snap["spot_proxy"],
        forward_proxy=snap["forward_proxy"],
        commodity_type_stats=[
            CommodityTypeEffectiveness(
                commodity_type=row["commodity_type"],
                modeled_count=row["modeled_count"],
                avg_hit_rate=row["avg_hit_rate"],
                avg_expected_return_net_cost=row["avg_expected_return_net_cost"],
                avg_score=row["avg_score"],
                top_bucket_avg_net_return=row["top_bucket_avg_net_return"],
                contract_coverage_pct=row.get("contract_coverage_pct", 0.0),
                avg_commodity_beta=row.get("avg_commodity_beta", 0.0),
                best_ticker=row.get("best_ticker", ""),
                best_hit_rate=row.get("best_hit_rate", 0.0),
            )
            for row in snap["commodity_type_stats"]
        ],
        opportunities=[
            GlobalOpportunity(
                ticker=row["ticker"],
                commodity_type=row["commodity_type"],
                country=row["country"],
                sector=row["sector"],
                direction=row["direction"],
                score=row["score"],
                spot_price=row["spot_price"],
                fair_value_price=row["fair_value_price"],
                expected_return_gross=row["expected_return_gross"],
                expected_return_net_cost=row["expected_return_net_cost"],
                cost_bps=row["cost_bps"],
                hit_rate=row["hit_rate"],
                mae=row["mae"],
                confidence=row["confidence"],
                predicted_margin_next=row["predicted_margin_next"],
                predicted_margin_change=row["predicted_margin_change"],
                production_growth_assumption=row["production_growth_assumption"],
                oil_beta=row["oil_beta"],
                oil_gamma=row["oil_gamma"],
                shipping_beta=row["shipping_beta"],
                shipping_gamma=row["shipping_gamma"],
                event_beta=row["event_beta"],
                event_gamma=row["event_gamma"],
                commodity_beta=row.get("commodity_beta", 0.0),
                market_cap=row["market_cap"],
                avg_daily_volume=row["avg_daily_volume"],
                risk_flags=row["risk_flags"],
                top_predictive_contracts=row["top_predictive_contracts"],
                top_features=row.get("top_features", []),
            )
            for row in snap["opportunities"]
        ],
    )


@router.get("/companies/{ticker}/narrative")
def get_ticker_narrative(ticker: str):
    """T1: Plain-English signal narrative for a ticker, derived from Ridge model feature importances."""
    snap = global_scan.scan_async()
    if snap is None:
        raise HTTPException(status_code=503, detail="computing: global scan in progress - retry in 2-3 minutes")
    ticker_upper = ticker.upper()
    match = next((o for o in snap.get("opportunities", []) if o["ticker"] == ticker_upper), None)
    if match is None:
        raise HTTPException(
            status_code=404,
            detail=f"{ticker_upper} not found in latest global scan. It may not be in the universe or had insufficient history.",
        )
    return build_narrative(match)


@router.post("/pipeline/run-daily")
def run_daily_pipeline(db: Session = Depends(get_db)):
    return orchestrator.run_daily(db)


@router.post("/pipeline/run-event-triggered")
def run_event_triggered_pipeline(db: Session = Depends(get_db)):
    return orchestrator.run_event_triggered(db)


@router.post("/pipeline/run-quarterly-fundamentals")
def run_quarterly_fundamentals(db: Session = Depends(get_db)):
    return orchestrator.run_quarterly_fundamentals(db)


@router.post("/events/news")
def receive_news_event(payload: dict, db: Session = Depends(get_db)):
    """
    T3: Receive a classified news event from the news pre-emption pipeline.
    Triggers event pipeline if impact_score >= 3.
    """
    event_type = payload.get("event_type", "")
    impact_score = payload.get("impact_score", 0)
    headline = payload.get("headline", "")

    if not event_type:
        raise HTTPException(status_code=400, detail="event_type is required")

    result = {
        "received": True,
        "event_type": event_type,
        "impact_score": impact_score,
        "headline": headline,
        "pipeline_triggered": False,
    }

    # Trigger event pipeline for high-impact events
    if impact_score >= 3:
        try:
            orchestrator.run_event_triggered(db)
            result["pipeline_triggered"] = True
        except Exception as exc:
            result["pipeline_error"] = str(exc)

    return result


@router.get("/news/monitor-status")
def get_news_monitor_status():
    """T3: Check the status of the background news monitor."""
    from app.providers.news_provider import get_or_start_monitor
    monitor = get_or_start_monitor()
    return monitor.status


